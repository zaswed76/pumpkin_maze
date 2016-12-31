#-----------------Импорт модулей-----------------
from tkinter import *
from tkinter import ttk
import tkinter.filedialog

import datetime,os

import openpyxl, difflib

from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import colors

from collections import Counter

from fuzzywuzzy import fuzz
from fuzzywuzzy import process


#-----------------Определение функций-----------------

# Диалоговое окно открытия файлов спецификаций
def LoadFileSO(ev):
    files = tkinter.filedialog.askopenfilenames (parent=root,title='Выбрать спецификации',filetypes = [("Excel files", "*.xlsx")])
    FileList = LIST1.get(0,END)
    for x in files:
        if (str(x)) not in FileList: LIST1.insert (END, str(x))
#----------Конец функции----------


# Диалоговое окно открытия файлов смет
def LoadFileSM(ev):
    files = tkinter.filedialog.askopenfilenames (parent=root,title='Выбрать сметы',filetypes = [("Excel files", "*.xlsx")])
    FileList = LIST2.get(0,END)
    for x in files:
        if (str(x)) not in FileList: LIST2.insert (END, str(x))
#----------Конец функции----------

        
# Удалить из списка выделенную спецификацию
def DelCurSelSO(ev):
    if LIST1.size() == 0: return
    z = LIST1.curselection()
    if len(z) == 0: return
    LIST1.delete(z)
#----------Конец функции----------


# Удалить из списка выделенную смету
def DelCurSelSM(ev):
    if LIST2.size() == 0: return
    z = LIST2.curselection()
    if len(z) == 0: return
    LIST2.delete(z)
#----------Конец функции----------


# Функция дизактивации кнопок
def ButtonOff():
    BUT1.config(state=DISABLED)
    BUT1.unbind("<Button-1>", None)
    BUT1.update()

    BUT2.config(state=DISABLED)
    BUT2.unbind("<Button-1>", None)
    BUT2.update()

    BUT3.config(state=DISABLED)
    BUT3.unbind("<Button-1>", None)
    BUT3.update()

    BUT4.config(state=DISABLED)
    BUT4.unbind("<Button-1>", None)
    BUT4.update()

    BUT5.config(state=DISABLED)
    BUT5.unbind("<Button-1>", None)
    BUT5.update()
#----------Конец функции----------


# Функция активации кнопок
def ButtonOn():
    BUT1.config(state=NORMAL)
    BUT1.bind("<Button-1>", LoadFileSO)
    BUT1.update()

    BUT2.config(state=NORMAL)
    BUT2.bind("<Button-1>", LoadFileSM)
    BUT2.update()

    BUT3.config(state=NORMAL)
    BUT3.bind("<Button-1>", DelCurSelSM)
    BUT3.update()

    BUT4.config(state=NORMAL)
    BUT4.bind("<Button-1>", Analyse)
    BUT4.update()

    BUT5.config(state=NORMAL)
    BUT5.bind("<Button-1>", DelCurSelSO)
    BUT5.update()
#----------Конец функции----------


# Функция запуска анализа
def Analyse(ev):
    ButtonOff()
    ListOfSO = LIST1.get(0,END) # Получить список спецификаций
    ListOfSM = LIST2.get(0,END) # Получить список смет
    if (len(ListOfSO) == 0) or (len(ListOfSM) == 0):
        LB3.config(text = "Спецификации или сметы не выбраны.")
        LB3.update()
        ButtonOn()
        return
    else:
        LB3.config(text = "Идет формирование базы данных")
        LB3.update()

        GlobalSOTable = FormGlobalTableSO(ListOfSO)        
        if GlobalSOTable == False:
            ButtonOn()
            return

        GlobalSMTable = FormGlobalTableSM(ListOfSM)
        if GlobalSMTable == False:
            ButtonOn()
            return

        GlobalData = FormGlobalTable(GlobalSOTable,GlobalSMTable)
        if GlobalData == False:
            ButtonOn()
            return
        else:
            LB3.config(text = "Идет запись в выходной файл tableout.xlsx")
            LB3.update()
            if WriteToFile(GlobalData) == False:
                ButtonOn()
                return
            LB3.config(text = "Записано в файл на рабочем столе")
            LB3.update()

    ButtonOn()
#----------Конец функции----------


# Функция формирования глобальной таблицы данных из спецификаций
def FormGlobalTableSO (ListOfSO):
    GlobalSOTable = []      #Глобальная таблица со всеми записями из всех выбранных спецификаций

    SOColumnName = 2        #Номер столбца "Наименование и техническая характеристика"
    SOColumnType = 3        #Номер столбца "Тип, марка, обозначение документа, опросного листа"
    SOColumnEI = 6          #Номер столбца "Ед. измерения"
    SOColumnQuantity = 7    #Номер столбца "Кол."
    SOColumnWeight = 8      #Номер столбца "Масса 1 ед., кг"

    #Начало цикла для формирования Глобальной таблицы данных из всех спецификаций GlobalSOTable
    #Структура элемента GlobalSOTable - [0-номер строки в спецификации, 1-наименование позиции, 2-тип позиции, 3-единица измерения, 4-количество, 5-вес одной единицы, 6-имя файла спецификации]
    for SOFileItem in ListOfSO:
        NameOfFile = str(SOFileItem)
        NameOfFile = NameOfFile.split("/")
        NameOfFile = NameOfFile [len (NameOfFile) - 1]

        LB3.config(text = "Формирование базы МТР из спецификации " + NameOfFile)
        LB3.update()

        try:
            SOBook = openpyxl.load_workbook(SOFileItem, data_only=True)
        except:
            LB3.config(text = "Ошибка открытия спецификации " + NameOfFile)
            LB3.update()
            return False

        try:
            SOSheet = SOBook['Ввод данных']
        except:
            LB3.config(text = "В спецификации "+NameOfFile+" отсутствует вкладка 'Ввод данных'")
            LB3.update()
            return False

        MaxRow = SOSheet.max_row
        MaxColumn = openpyxl.cell.get_column_letter(SOSheet.max_column)        
        CellData = []
        for RowItem in range (4, MaxRow+1):
            CellRange = tuple(SOSheet["A"+str(RowItem)+":"+MaxColumn+str(RowItem)])
            if "Искл" in str(CellRange[0][9].value): continue
            CellData = []
            for CellRow in CellRange:                
                for Cell in CellRow:
                    CellData.append (Cell.value)
            if all (x == None for x in CellData):
                continue
            else:
                if CellData[SOColumnName] == None: CellData[SOColumnName] = ""
                GlobalSOTable.append([RowItem,str(CellData[SOColumnName]),str(CellData[SOColumnType]),CellData[SOColumnEI],CellData[SOColumnQuantity],CellData[SOColumnWeight],NameOfFile])

    for QYCounter in range (0,len(GlobalSOTable)):
        if GlobalSOTable[QYCounter][4] == None:
            continue
        else:
            QYNorm = str(GlobalSOTable[QYCounter][4]).split()
            GlobalSOTable[QYCounter][4] = QYNorm[len(QYNorm)-1].replace(".",",")

    LIST1.delete(0,END)

    return GlobalSOTable
#----------Конец функции----------


# Функция формирования глобальной таблицы данных из смет
def FormGlobalTableSM (ListOfSM):
    SMMarks = ("ОЗП=","ЭМ=","ЗПМ=","МАТ=")  #Ключевые слова по которым определяем объем
    SMColumnPos = 0                         #Номер столбца "№ пп"
    SMColumnName = 2                        #Номер столбца "Наименование"
    SMColumnEI = 3                          #Номер столбца "Ед. изм."
    SMColumnQuantity = 4                    #Номер столбца "Кол."
    SMColumnPrice = 5                       #Номер столбца "Стоимость единицы, руб. Всего"

    GlobalSMTable = [] #Глобальная таблица с расценками объемов

    #Структура GlobalSMTable - [номер строки в смете,позиция в смете,наименование расценки,ед.изм.,кол-во,цена единицы,стоимость,имя файла сметы]

    #Начало цикла в котором формируется Глобальная таблица с расценками объемов для каждой отдельной сметы
    for SMFileItem in ListOfSM:
        NameOfFile = str(SMFileItem)
        NameOfFile = NameOfFile.split("/")
        NameOfFile = NameOfFile [len (NameOfFile) - 1]

        LB3.config(text = "Формирование базы расценок из сметы " + NameOfFile)
        LB3.update()

        try:
            SMBook = openpyxl.load_workbook(SMFileItem, data_only=True)
        except:
            LB3.config(text = "Ошибка открытия сметы " + NameOfFile)
            LB3.update()
            return False

        SMSheet = SMBook.get_sheet_names()
        for x in SMSheet :
            if "ЛСР" in x :
                SMSheet = SMBook[str(x)]
                break

        try:
            MaxRow = SMSheet.max_row
        except:
            LB3.config(text = "В смете " + NameOfFile + " отсутсвует вкладка с наименованием 'ЛСР'")
            LB3.update()
            return False            

        MaxColumn = openpyxl.cell.get_column_letter(SMSheet.max_column)
        CellData = []
        for RowItem in range (1, MaxRow+1):
            CellRange = tuple(SMSheet["A"+str(RowItem)+":"+MaxColumn+str(RowItem)])
            for CellRow in CellRange :        
                for Cell in CellRow :
                    CellData.append (Cell.value)
            if all (x == None for x in CellData):
                CellData = []
            else:
                if (all (x in str(CellData[SMColumnName]) for x in SMMarks)) and (" ФОТ" not in CellData[SMColumnName]):
                    GlobalSMTable.append([RowItem,CellData[0],CellData[2],CellData[3],CellData[4],CellData[5],CellData[9],NameOfFile])
                elif ("ПЗ=" in str(CellData[SMColumnName])) and (" ФОТ" not in CellData[SMColumnName]):
                    GlobalSMTable.append([RowItem,CellData[0],CellData[2],CellData[3],CellData[4],CellData[5],CellData[9],NameOfFile])
                elif ("МАТ=" in str(CellData[SMColumnName])) and (" ФОТ" not in CellData[SMColumnName]):
                    GlobalSMTable.append([RowItem,CellData[0],CellData[2],CellData[3],CellData[4],CellData[5],CellData[9],NameOfFile])
                elif ("МАТ=" in str(CellData[SMColumnName])) and (all (x not in str(CellData[SMColumnName]) for x in SMMarks)):
                    GlobalSMTable.append([RowItem,CellData[0],CellData[2],CellData[3],CellData[4],CellData[5],CellData[9],NameOfFile])
                CellData = []

    for SMCounter in range (0,len(GlobalSMTable)):
        NameNormalize = ClearName(GlobalSMTable[SMCounter][2])
        GlobalSMTable[SMCounter][2] = NameNormalize
        GlobalSMTable[SMCounter].append(QYNormalize(GlobalSMTable[SMCounter][4]).replace(".",","))

    LIST2.delete(0,END)
    
    return GlobalSMTable
#----------Конец функции----------


#Функция нормализации количества МТР из сметного расчета (приведение к ед.изм. спецификации)
def QYNormalize(QYInput):
    QYOut = str(QYInput).split("\n")
    QYOut = QYOut[len(QYOut)-1]
    
    for QYSign in QYOut:
        if QYSign in ("*","/"): #,"+","-"):
            QYOut = QYOut.split(QYSign)
            QYOut = QYOut[0]
            QYOut = QYOut.strip()
            break    
    return QYOut
#----------Конец функции----------


#Функция очистки Наименования сметной расценки от лишних значений (стоимости, коэффициентов)
def ClearName (Name):
    LeftMarks = ("Цена","ц/о","ЦО","ИНДЕКС","Индекс","МАТ=","ПЗ=",":")
    for Mark in LeftMarks:
        if Mark in Name:
            NameOut = Name.split(Mark)[0]
            break
        else:
            NameOut = Name
    return NameOut
#----------Конец функции----------


#Функция сравнения по словам, на выходе процент
def CompareWordPercent (SOItem, SMItem):
    SOName = SOItem[1]
    if SOItem[2] != None:
        SOName = SOName+" "+SOItem[2]
    Percent = difflib.SequenceMatcher (None, SOName, SMItem[2]).ratio() * 100
    return Percent
#----------Конец функции----------


#Функция сравнения по буквам, на выходе процент
def CompareLetterPercent (SOItem, SMItem):
    SOName = SOItem[1]
    if SOItem[2] != None:
        SOName = SOName+" "+SOItem[2]
    SOLetterCounter = Counter (SOName)
    SMLetterCounter = Counter (SMItem[2])
    Percent = difflib.SequenceMatcher (None, SMLetterCounter, SOLetterCounter).ratio() * 100
    return Percent
#----------Конец функции----------


#Функция сравнения FuzzyWuzzy
def CompareFuzzy (SOItem, SMItem):
    SOName = SOItem[1]
    if SOItem[2] != None:
        SOName = SOName+" "+SOItem[2]
    Percent = fuzz.QRatio (SOName, SMItem[2])
    return Percent
#----------Конец функции----------


#Функция формирования глобальной талицы с данными
def FormGlobalTable (GlobalSOTable, GlobalSMTable):

    GlobalDataTable = []    #Выходная таблица для записи в Эксель файл
    PercentMatrix = []      #Матрица содержащая все позиции по спецификации, с процентами совпадения для всех расценок для каждой позиции спецификации
    
    #Цикл для формирования Глобальной таблицы данных PercentMatrix (отдельная для каждой сметы) со сравнением СО и СМ
    #Структура PercentMatrix - [Элементы SOItemPercent ...]
    #Структура SOItemPercent - [Элемент из GlobalSOTable, Элементы SMItemsPercent отсортированные в порядке убывания процента]
    #Структура SMItemsPercent - [Элемент из GlobalSMTable, результат сравнения в виде процента, флаг совпадения количества]

    for SOItem in GlobalSOTable:
        ProgressText = "Формирование таблиц соответствия. Выполнено "+str(GlobalSOTable.index(SOItem))+" из "+str(len(GlobalSOTable))
        LB3.config(text = ProgressText)
        LB3.update()

        SOItemPercent = []      #Очищаем списки
        SMItemsPercent = []     #Очищаем списки

        if SOItem[4] == None: continue      #Если значение количества для элемента спецификации отсутсвует (None), то пропускаем данный элемент (пропуск итерации цикла)

        SOItemPercent.append(SOItem)
        
        for SMItem in GlobalSMTable:
            QYCompareStatus = False
            TotalPercent = round(CompareWordPercent(SOItem,SMItem),10) + round(CompareWordPercent(SOItem,SMItem),10) + CompareFuzzy (SOItem,SMItem)
            if SOItem[4] == SMItem[8]:
                TotalPercent += 100
                QYCompareStatus = True
            SMItemsPercent.append((SMItem,TotalPercent,QYCompareStatus))

        SMItemsPercent.sort(key=lambda x: x[1],reverse=True)            
        SOItemPercent.append(SMItemsPercent)
        PercentMatrix.append(SOItemPercent)

    #Цикл распределения расценок по позициям из смет в зависимости от процента
    PMPosition = 0
    PMLenght = len (PercentMatrix) - 1

    while (PMPosition != PMLenght+1):
        ProgressText = "Процесс распределения расценок. Выполнено "+str(PMPosition)+" из "+str(PMLenght)
        LB3.config(text = ProgressText)
        LB3.update()

        ChangesExist = False

        PMItemFirst = PercentMatrix[PMPosition]
        if len(PMItemFirst[1]) == 0:
                PMPosition += 1
                continue
        for PMCounter in range (PMPosition + 1, PMLenght + 1):
            PMItemNext = PercentMatrix[PMCounter]
            if len(PMItemNext[1]) == 0 :
                continue
            for SMItemCounter in PMItemNext[1]:
                if PMItemFirst[1][0][0][0] == SMItemCounter[0][0]:
                    if PMItemFirst[1][0][1] < SMItemCounter[1]:
                        if PMItemNext[1].index(SMItemCounter) == 0:
                            PercentMatrix[PMPosition][1].remove(PercentMatrix[PMPosition][1][0])                    
                            ChangesExist = True
                        else:
                            PercentMatrix[PMCounter][1].remove(SMItemCounter)
                    else:
                        PercentMatrix[PMCounter][1].remove(SMItemCounter)
                       
                    break

            if ChangesExist == True :
                break
                
        if ChangesExist == False:
            PercentMatrix[PMPosition][1] = [PMItemFirst[1][0]]
            PMItemFirst[1] = [PMItemFirst[1][0]]
            PMPosition += 1


    #Структура элемента GlobalSOTable - [0-номер строки в спецификации, 1-наименование позиции, 2-тип позиции, 3-единица измерения, 4-количество, 5-вес одной единицы, 6-имя файла спецификации]
    for SOItem in GlobalSOTable:
        if SOItem[4] == None:
            GlobalDataTable.append([SOItem[1],SOItem[2],SOItem[3],SOItem[4],SOItem[6],SOItem[0],None,None,None,None,None,None,None,None,None,None,None,None])
        else:
            for PMItem in PercentMatrix:
                if PMItem[0][0] == SOItem[0]:
                    if len(PMItem[1]) == 0:
                        GlobalDataTable.append([SOItem[1],SOItem[2],SOItem[3],SOItem[4],SOItem[6],SOItem[0],None,None,None,None,None,None,None,None,None,None,None,False])
                    else:
                        LocalSMName = PMItem[1][0][0][2]
                        LocalSMEI = PMItem[1][0][0][3]
                        LocalSMQY = str(PMItem[1][0][0][4])
                        LocalSMPrice = str(PMItem[1][0][0][5]).split()[0]
                        LocalSMValue = str(PMItem[1][0][0][6])
                        LocalSMPercent = round(PMItem[1][0][1],3)
                        LocalSMNumber = PMItem[1][0][0][0]
                        LocalSMFileName = PMItem[1][0][0][7]
                        GlobalDataTable.append([SOItem[1],SOItem[2],SOItem[3],SOItem[4],SOItem[6],SOItem[0],LocalSMName,LocalSMEI,LocalSMQY,LocalSMFileName,LocalSMNumber,LocalSMPercent,None,None,LocalSMValue,LocalSMPrice,None,PMItem[1][0][2]])
                        break

    return GlobalDataTable
#----------Конец функции----------


# Функция записи в выходной файл tableout.xlsx на рабочий стол
def WriteToFile(Data):
    TableName = 'tablein.xlsx'
    date = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")
    TableOut = os.path.join(os.path.expanduser("~"), "Desktop")+'\\'+date+'_'+'tableout.xlsx'
    TableRow = 6

    try:
        TableBook = openpyxl.load_workbook(TableName)
    except:
        LB3.config(text = "Ошибка открытия файла tablein.xlsx")
        LB3.update()
        return False

    TableSheet = TableBook['Входная']

    for DataItem in Data:
        TableSheet["A"+str(TableRow)].value = DataItem[0]   #Наименование по спецификации
        TableSheet["B"+str(TableRow)].value = DataItem[1]   #Тип по спецификации
        TableSheet["C"+str(TableRow)].value = DataItem[2]   #Единица измерения
        TableSheet["D"+str(TableRow)].value = DataItem[3]   #Количество по спецификации
        TableSheet["E"+str(TableRow)].value = DataItem[4]   #Имя файла спецификации
        TableSheet["F"+str(TableRow)].value = DataItem[5]   #Номер строки файла Excel спецификации
        TableSheet["G"+str(TableRow)].value = DataItem[6]   #Наименование по сметам

        if (DataItem[3] != None) and (DataItem[6] == None) and (DataItem[8] == None):
            TableSheet["G"+str(TableRow)].fill = PatternFill (fill_type="solid", start_color=colors.RED, end_color=colors.DARKRED)
            TableSheet["I"+str(TableRow)].fill = PatternFill (fill_type="solid", start_color=colors.RED, end_color=colors.DARKRED)
        elif (DataItem[3] != None) and (DataItem[11] != None) and (DataItem[8] != None):
            if DataItem[11]<100 :
                TableSheet["G"+str(TableRow)].fill = PatternFill (fill_type="solid", start_color=colors.RED, end_color=colors.RED)
            elif 150>DataItem[11]>=100 :
                TableSheet["G"+str(TableRow)].fill = PatternFill (fill_type="solid", start_color=colors.YELLOW, end_color=colors.YELLOW)

        TableSheet["H"+str(TableRow)].value = DataItem[7]   #Единица измерения
        TableSheet["I"+str(TableRow)].value = DataItem[8]   #Количество по сметам

        if DataItem[17] == False:
            ChangeFontColor = TableSheet["I"+str(TableRow)]
            FontColor = Font(color=colors.RED)
            ChangeFontColor.font = FontColor

        TableSheet["J"+str(TableRow)].value = DataItem[9]   #Имя файла сметы
        TableSheet["K"+str(TableRow)].value = DataItem[10]  #Номер строки файла Excel сметы
        TableSheet["L"+str(TableRow)].value = DataItem[11]  #Примечание (записывается процент совпадения)
        TableSheet["M"+str(TableRow)].value = DataItem[12]  #Производство (Отечественное или импортное, проставляется всегда None)
        TableSheet["N"+str(TableRow)].value = DataItem[13]  #Аналог (проставляется всегда None)
        TableSheet["O"+str(TableRow)].value = DataItem[14]  #Стоимость общая (= Количество по сметам * Цена за единицу) (пока прописывается None)
        TableSheet["P"+str(TableRow)].value = DataItem[15]  #Цена за единицу
        TableSheet["Q"+str(TableRow)].value = DataItem[16]  #НВЛ (проставляется всегда None)
        TableRow += 1
    TableRow = 6

    try:
        TableBook.save(TableOut)
    except:
        LB3.config(text = "Ошибка сохранения в файл "+TableOut)
        LB3.update()
        return False        

    return
#----------Конец функции----------


# Запуск основной программы
SetUserFont = "Calibri 12"

root = Tk()

x = (root.winfo_screenwidth() - 650) / 2
y = (root.winfo_screenheight() - 600) / 2
root.wm_geometry("+%d+%d" % (x, y))

LB1 = Label (root, text="Список спецификаций для проверки", font=SetUserFont)
LB1.grid(row=0,column=1)

LIST1 = Listbox(root,font=(SetUserFont),height=5,width=50)
LIST1.grid(row=1,column=1)

BUT1 = ttk.Button(root)
BUT1["text"] = "Выбрать"
BUT1.bind("<Button-1>", LoadFileSO)
BUT1.grid(row=1, column=2, sticky="n")

BUT5 = ttk.Button(root)
BUT5["text"] = "Исключить"
BUT5.bind("<Button-1>", DelCurSelSO)
BUT5.grid(row=1, column=2)

LB2 = Label (root, text="Список смет для проверки", font=SetUserFont)
LB2.grid(row=2,column=1)

LIST2 = Listbox(root,font=(SetUserFont),height=5,width=50)
LIST2.grid(row=3,column=1)

BUT2 = ttk.Button(root)
BUT2["text"] = "Выбрать"
BUT2.bind("<Button-1>", LoadFileSM)
BUT2.grid(row=3, column=2, sticky="n")

BUT3 = ttk.Button(root)
BUT3["text"] = "Исключить"
BUT3.bind("<Button-1>", DelCurSelSM)
BUT3.grid(row=3, column=2)

BUT4 = ttk.Button(root, width=20)
BUT4["text"] = "Анализ"
BUT4.bind("<Button-1>", Analyse)
BUT4.grid(row=4, column=1)

LB3 = Label (root, text = "Информационные сообщения", font=SetUserFont)
LB3.grid(row=5,column=1)

root.mainloop()
